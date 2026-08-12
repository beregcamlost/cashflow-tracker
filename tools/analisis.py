#!/usr/bin/env python3
"""Análisis financiero del Cashflow sheet exportado a xlsx.

Uso: python analisis.py <ruta a Cashflow.xlsx>
Lee las pestañas del tracker, imprime un informe a stdout y verifica
la consistencia de CUOTAS (stats del header vs tabla, planes stale,
posibles duplicados por múltiples statements del mismo plan).
"""
import re
import sys
from datetime import date, datetime, timedelta

import openpyxl
import pandas as pd

CUOTA_RE = re.compile(r"(?:CC|CF)\s+(\d{2})-(\d{2})", re.I)
EXCLUDE_CATS = {"CC Payment", "Transfers Out", "Transfers In"}


def find_sheet(wb, name):
    for ws in wb.worksheets:
        if ws.title.split()[-1] == name or name in ws.title:
            return ws
    return None


def sheet_df(ws, header_row=1):
    rows = list(ws.iter_rows(min_row=header_row, values_only=True))
    if not rows:
        return pd.DataFrame()
    header = [str(c) if c is not None else "" for c in rows[0]]
    data = [r for r in rows[1:] if any(c is not None and c != "" for c in r)]
    return pd.DataFrame(data, columns=header)


def fmt(n):
    return f"{n:,.0f}".replace(",", ".")


def months_elapsed(d, today):
    if isinstance(d, datetime):
        d = d.date()
    if not isinstance(d, date):
        return 0
    m = (today.year - d.year) * 12 + (today.month - d.month)
    if today.day < d.day:
        m -= 1
    return max(0, m)


def config_value(cfg, key):
    row = cfg[cfg.iloc[:, 0].astype(str).str.strip().str.upper() == key]
    if row.empty:
        return None
    v = row.iloc[0, 1]
    return float(v) if isinstance(v, (int, float)) else None


def report_movements(df, label, amount_col, negative=False):
    df = df.copy()
    df["Fecha"] = pd.to_datetime(df["Fecha"], errors="coerce")
    df[amount_col] = pd.to_numeric(df[amount_col], errors="coerce")
    df = df.dropna(subset=["Fecha", amount_col])
    if df.empty:
        print(f"\n== {label}: sin datos ==")
        return None
    print(f"\n== {label} — {len(df)} movimientos, "
          f"{df['Fecha'].min():%Y-%m-%d} → {df['Fecha'].max():%Y-%m-%d} ==")
    if negative:
        df = df[df[amount_col] < 0].copy()
        df[amount_col] = -df[amount_col]
    spend = df[(df[amount_col] > 0) & ~df["Categoria"].isin(EXCLUDE_CATS)]
    monthly = spend.groupby(df["Fecha"].dt.to_period("M"))[amount_col].sum()
    print("Gasto por mes (sin pagos TC/transferencias):")
    for period, total in monthly.items():
        print(f"  {period}  {fmt(total):>12}")
    top = (spend.groupby("Categoria")[amount_col].sum()
           .sort_values(ascending=False).head(10))
    print("Top categorías (todo el período):")
    for cat, total in top.items():
        print(f"  {cat:<20} {fmt(total):>12}")
    return spend


def report_deuda_tc(wb):
    """Espejo de computeCycleDebt_ del tracker: statement del último cierre,
    pagos posteriores y por facturar — para auditar la cifra del dashboard."""
    usd = 950.0
    cfg_ws = find_sheet(wb, "CONFIG")
    if cfg_ws is not None:
        usd = config_value(sheet_df(cfg_ws), "USD_CLP") or usd
    print("\n== DEUDA TC HOY (est.) ==")
    total = 0
    for name, factor in [("MOV_CC_NACIONAL", 1.0), ("MOV_CC_INTL", usd)]:
        ws = find_sheet(wb, name)
        if ws is None:
            continue
        rows = [r for r in ws.iter_rows(min_row=2, values_only=True)
                if r[0] is not None]
        facts = [r[0] for r in rows if str(r[3]).strip() == "Facturado"
                 and isinstance(r[2], (int, float)) and r[2] > 0]
        cierre = max(facts) if facts else None
        st = pp = pf = 0
        for f, _, m, t, *_ in rows:
            if not isinstance(m, (int, float)):
                continue
            t = str(t).strip()
            if cierre is None:
                if t == "No Facturado" and m > 0:
                    pf += m
                continue
            if (t == "Facturado" and m > 0 and f <= cierre
                    and (cierre - f) <= timedelta(days=31)):
                st += m
            elif m < 0 and f > cierre:
                pp += -m
            elif t == "No Facturado" and m > 0 and f > cierre:
                pf += m
        st, pp, pf = st * factor, pp * factor, pf * factor
        deuda = max(0, st - pp) + pf
        total += deuda
        print(f"  {name}: cierre {cierre.date() if cierre else '—'}  "
              f"statement {fmt(st)}  pagos post {fmt(pp)}  "
              f"por facturar {fmt(pf)}  → deuda {fmt(deuda)}")
    print(f"  TOTAL deuda TC hoy: {fmt(total)}")
    return total


MESES = {"enero": 1, "febrero": 2, "marzo": 3, "abril": 4, "mayo": 5,
         "junio": 6, "julio": 7, "agosto": 8, "septiembre": 9,
         "octubre": 10, "noviembre": 11, "diciembre": 12}


def report_casa(wb, today):
    """Planilla compartida de la casa: reparto 60/40 por mes + compras grandes."""
    print("\n== GASTOS CASA (planilla Gordo) ==")
    shares = {}
    for ws in wb.worksheets:
        mes = MESES.get(ws.title.strip().lower().split()[-1])
        if mes is None:
            continue
        # Filas 1-2 de cada tab mensual: parte de la otra familia, parte propia
        vals = []
        for row in ws.iter_rows(max_row=5, values_only=True):
            if row[0] and isinstance(row[1], (int, float)):
                vals.append(row[1])
            if len(vals) == 2:
                break
        if len(vals) == 2:
            shares[mes] = vals[1]
            partial = "  (mes en curso, parcial)" if mes == today.month else ""
            print(f"  {today.year}-{mes:02d}  total casa {fmt(sum(vals)):>12}"
                  f"  tu parte {fmt(vals[1]):>12}{partial}")
    casa = {"shares": shares, "arriendo": None, "cuotas_mes": None, "total_mes": None}
    # Hoja "Gordo": col G = cuota mensual comprometida por las compras grandes;
    # L1 = total del mes; J2 = arriendo
    gordo = wb["Gordo"] if "Gordo" in wb.sheetnames else None
    if gordo is not None:
        items_cols = [(i, str(c.value).strip()) for i, c in enumerate(gordo[2])
                      if 2 <= i <= 5 and c.value and str(c.value).strip()]
        for row in gordo.iter_rows(min_row=3, values_only=True):
            d = row[0]
            if isinstance(d, datetime) and (d.year, d.month) == (today.year, today.month):
                items = [(h, row[i]) for i, h in items_cols
                         if isinstance(row[i], (int, float)) and row[i] > 0]
                cuota = row[6] if isinstance(row[6], (int, float)) and row[6] > 0 \
                    else sum(v for _, v in items)
                if cuota:
                    det = ", ".join(f"{h} {fmt(v)}" for h, v in items)
                    print(f"  Cuotas casa {today:%Y-%m}: {fmt(cuota)}  ({det})")
                    casa["cuotas_mes"] = cuota
                break
        arriendo = gordo.cell(2, 10).value
        total_mes = gordo.cell(1, 12).value
        if isinstance(arriendo, (int, float)):
            casa["arriendo"] = arriendo
            print(f"  Arriendo: {fmt(arriendo)}")
        if isinstance(total_mes, (int, float)):
            casa["total_mes"] = total_mes
            print(f"  Total del mes (L1): {fmt(total_mes)}  "
                  f"(arriendo + 40% vivienda + cuotas)")
    return casa


def verify_cuotas(wb, today):
    ws = find_sheet(wb, "CUOTAS")
    if ws is None:
        print("\n== CUOTAS: pestaña no encontrada ==")
        return
    header = {
        "active": ws.cell(2, 2).value,
        "monthly": ws.cell(2, 7).value,
        "remaining": ws.cell(3, 2).value,
        "finishing": ws.cell(3, 7).value,
    }
    rows = []
    for r in ws.iter_rows(min_row=6, values_only=True):
        if r[0] is None:
            break
        rows.append(dict(zip(
            ["fuente", "fecha", "desc", "monto", "moneda",
             "cuota", "total", "restantes", "rest_est"], r[:9])))
    print(f"\n== CUOTAS — {len(rows)} planes en tabla ==")
    if not rows:
        return

    usd_clp = 950.0
    cfg_ws = find_sheet(wb, "CONFIG")
    if cfg_ws is not None:
        cfg = sheet_df(cfg_ws)
        usd_clp = config_value(cfg, "USD_CLP") or usd_clp

    def clp(r, field):
        v = r[field] or 0
        return v * usd_clp if r["moneda"] == "USD" else v

    active = [r for r in rows if (r["restantes"] or 0) > 0]
    calc = {
        "active": len(active),
        "monthly": sum(clp(r, "monto") for r in active),
        "remaining": sum(clp(r, "rest_est") for r in active),
        "finishing": sum(1 for r in rows if r["restantes"] == 1),
    }
    labels = {"active": "Planes activos", "monthly": "Pago mensual",
              "remaining": "Total restante", "finishing": "Terminan próximo mes"}
    print(f"{'':<22}{'header':>12}{'recalculado':>12}")
    mismatch = False
    for k, lbl in labels.items():
        h, c = header[k] or 0, calc[k]
        flag = "" if abs(h - c) < 1 else "  <-- NO CUADRA"
        mismatch = mismatch or bool(flag)
        print(f"{lbl:<22}{fmt(h):>12}{fmt(c):>12}{flag}")
    if mismatch:
        print("FLAG: el header del sheet difiere del recálculo "
              "(¿stats contando planes terminados?).")

    stale = [r for r in active
             if months_elapsed(r["fecha"], today) >= (r["restantes"] or 0) > 0]
    if stale:
        print("FLAG planes stale (por fecha ya deberían estar terminados; "
              "faltan statements por importar):")
        for r in stale:
            print(f"  {str(r['desc'])[:40]:<40} cuota {r['cuota']:.0f}/"
                  f"{r['total']:.0f} visto {str(r['fecha'])[:10]}")

    seen = {}
    for r in rows:
        key = (CUOTA_RE.sub("", str(r["desc"])).strip(),
               r["total"], round(r["monto"] or 0))
        seen.setdefault(key, []).append(r)
    dups = {k: v for k, v in seen.items() if len(v) > 1}
    if dups:
        print("FLAG posibles duplicados (mismo plan en varios statements — "
              "se contaría doble en las stats):")
        for k, v in dups.items():
            pos = ", ".join(f"{r['cuota']:.0f}/{r['total']:.0f}" for r in v)
            print(f"  {k[0][:40]:<40} posiciones: {pos}")

    proj_active = [r for r in active
                   if months_elapsed(r["fecha"], today) < (r["restantes"] or 0)]
    proj_monthly = sum(clp(r, "monto") for r in proj_active)
    proj_rem = sum(clp(r, "monto") * ((r["restantes"] or 0)
                   - months_elapsed(r["fecha"], today)) for r in proj_active)
    print(f"Proyección a hoy (descontando meses transcurridos): "
          f"{len(proj_active)} planes, {fmt(proj_monthly)}/mes, "
          f"restante {fmt(proj_rem)}")


def main():
    path = sys.argv[1] if len(sys.argv) > 1 else "out/Cashflow.xlsx"
    casa_path = sys.argv[2] if len(sys.argv) > 2 else None
    today = date.today()
    wb = openpyxl.load_workbook(path, data_only=True)
    print(f"Cashflow — informe {today:%Y-%m-%d} ({path})")

    cfg_ws = find_sheet(wb, "CONFIG")
    salario = None
    if cfg_ws is not None:
        cfg = sheet_df(cfg_ws)
        salario = config_value(cfg, "SALARIO")
        print("\n== CONFIG ==")
        for key in ["SALARIO", "HOUSING", "FAMILIA", "USD_CLP"]:
            v = config_value(cfg, key)
            print(f"  {key:<10} {fmt(v) if v is not None else '—'}")
    if salario is None:
        print("  SALARIO no legible en CONFIG — corre Refresh en el sheet; "
              "sin salario no se calcula el margen")

    spends = []
    for name, col, neg in [("MOV_CC_NACIONAL", "Monto_CLP", False),
                           ("MOV_CC_INTL", "Monto_CLP_est", False),
                           ("MOV_BANCO", "Monto_CLP", True)]:
        ws = find_sheet(wb, name)
        if ws is not None:
            s = report_movements(sheet_df(ws), name, col, negative=neg)
            if s is not None and name != "MOV_BANCO":
                spends.append(s.rename(columns={col: "monto"}))

    verify_cuotas(wb, today)
    deuda_tc = report_deuda_tc(wb)

    casa = {}
    if casa_path:
        try:
            casa = report_casa(
                openpyxl.load_workbook(casa_path, data_only=True), today) or {}
        except Exception as e:
            print(f"\n== GASTOS CASA: no se pudo leer {casa_path}: {e} ==")

    if salario and spends:
        allsp = pd.concat([s[["Fecha", "monto"]] for s in spends])
        monthly = allsp.groupby(allsp["Fecha"].dt.to_period("M"))["monto"].sum()
        complete = monthly.iloc[-4:-1] if len(monthly) > 4 else monthly
        avg = complete.mean()
        print(f"\n== MARGEN (TC nacional + intl, sin banco) ==")
        print(f"  Salario            {fmt(salario):>12}")
        print(f"  Gasto TC prom/mes  {fmt(avg):>12}  "
              f"(meses: {', '.join(str(p) for p in complete.index)})")
        casa_total = 0
        shares = casa.get("shares") or {}
        comp = [v for m, v in sorted(shares.items()) if m != today.month][-3:]
        if comp:
            share_avg = sum(comp) / len(comp)
            casa_total += share_avg
            print(f"  Casa (tu 40%) prom {fmt(share_avg):>12}  "
                  f"(últimos {len(comp)} meses completos)")
        if casa.get("arriendo"):
            casa_total += casa["arriendo"]
            print(f"  Arriendo           {fmt(casa['arriendo']):>12}")
        if casa.get("cuotas_mes"):
            casa_total += casa["cuotas_mes"]
            print(f"  Cuotas casa        {fmt(casa['cuotas_mes']):>12}")
        print(f"  Margen bruto       {fmt(salario - avg - casa_total):>12}")
        if casa_total:
            print("  Nota: ítems de casa pagados con tu TC se solapan con el gasto TC "
                  "— el margen real puede ser algo mayor.")

    dash_ws = find_sheet(wb, "DASHBOARD")
    if dash_ws is not None and cfg_ws is not None:
        saldo = dash_ws.cell(18, 2).value
        cfg = sheet_df(cfg_ws)
        payday = config_value(cfg, "PAYDAY")
        sueldo_pd = config_value(cfg, "SUELDO_PAYDAY")
        housing = config_value(cfg, "HOUSING") or 0
        familia = config_value(cfg, "FAMILIA") or 0
        if not sueldo_pd:
            print("\n== LIBRE HOY: agrega PAYDAY y SUELDO_PAYDAY al CONFIG "
                  "del sheet para activar esta sección ==")
        if isinstance(saldo, (int, float)) and sueldo_pd:
            banco_ws = find_sheet(wb, "MOV_BANCO")
            abono = False
            if banco_ws is not None:
                for r in banco_ws.iter_rows(min_row=2, values_only=True):
                    if (isinstance(r[2], (int, float)) and r[2] >= 0.8 * sueldo_pd
                            and isinstance(r[0], datetime)
                            and (r[0].year, r[0].month) == (today.year, today.month)):
                        abono = True
                        break
            pendiente = 0 if abono else sueldo_pd
            libre_hoy = saldo + pendiente - deuda_tc - housing - familia
            print(f"\n== LIBRE HOY (est.) ==")
            print(f"  Saldo banco        {fmt(saldo):>12}")
            print(f"  Sueldo por entrar  {fmt(pendiente):>12}"
                  + (f"  (día {payday:.0f})" if pendiente and payday else ""))
            print(f"  Deuda TC hoy      -{fmt(deuda_tc):>12}")
            print(f"  Casa (total mes)  -{fmt(housing):>12}  (asume no pagada aún)")
            print(f"  Familia           -{fmt(familia):>12}")
            print(f"  LIBRE HOY          {fmt(libre_hoy):>12}")


if __name__ == "__main__":
    main()
